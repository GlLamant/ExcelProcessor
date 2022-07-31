from openpyxl import load_workbook


class BatchLookupAndCopy:
    def start(self,
              src_file, src_id_col, src_copy_col, src_title_rows,
              dst_file, dst_id_col, dst_paste_col, dst_title_rows):
        """根据用户输入的控制参数开始批量查找、并复制

        :param src_file: 源文件路径
        :type src_file: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`

        :param src_id_col: 源文件中身份证号码所在列
        :type src_id_col: string, e.g. a, b, c

        :param src_copy_col: 源文件需要拷贝内容的所在列
        :type src_copy_col: string, e.g. a, b, c

        :param src_title_rows: 源文件标题行数
        :type src_title_rows: int

        :param dst_file: 目标文件路径
        :type dst_file: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`

        :param dst_id_col: 目标文件中身份证号码所在列
        :type dst_id_col: string, e.g. a, b, c

        :param dst_paste_col: 目标文件需要拷贝内容到该列
        :type dst_paste_col: string, e.g. a, b, c

        :param dst_title_rows: 目标文件标题行数
        :type dst_title_rows: int

        :rtype: :class:`openpyxl.workbook.Workbook`

        .. note::

            When using lazy load, all worksheets will be :class:`openpyxl.worksheet.iter_worksheet.IterableWorksheet`
            and the returned workbook will be read-only.

        """

        # 加载excel文件
        src_excel = load_workbook(src_file)
        dst_excel = load_workbook(dst_file)

        # 获取activa的sheet
        src_active_sheet = src_excel.active
        dst_active_sheet = dst_excel.active

        # src_title_rows + 1 的原因是防止有多行标题，导致目标excel标题行被改动
        for src_row in range(src_title_rows + 1, src_active_sheet.max_row + 1):
            src_id = (src_active_sheet[src_id_col + str(src_row)]).value
            for dst_row in range(dst_title_rows + 1, dst_active_sheet.max_row + 1):
                dst_id = (dst_active_sheet[dst_id_col + str(dst_row)]).value
                # 对比身份证号码
                if src_id == dst_id:
                    # 执行拷贝动作
                    dst_active_sheet[dst_paste_col + str(dst_row)].value = src_active_sheet[
                        src_copy_col + str(src_row)].value

        dst_excel.save("../../doc/final_excel.xlsx")


batch_loopup_and_copy = BatchLookupAndCopy()
batch_loopup_and_copy.start("../../doc/表1.xlsx", 'd', 'f', 1, "../../doc/表4.xlsx", 'd', 'g', 1)
