import tkinter
import tkinter.messagebox
from tkinter import *
from enum import Enum
from tkinter.filedialog import (askopenfilename,
                                askopenfilenames,
                                askdirectory,
                                asksaveasfilename)
from openpyxl import load_workbook


# 获取tk控件的大小
def get_tk_size(tk_instance):
    tk_instance.update()
    return tk_instance.winfo_width(), tk_instance.winfo_height()


# 两个excel文件的背景色
kRED = '#c75450'
kGREEN = '#4f7d58'


class FrameType(Enum):
    kUnknown = -1
    kSrcFrame = 0
    kDstFrame = 1


# tkinter.Frame的包装类，包含有若干个LabelFrame
class ExcelWrapper:
    __FONT = '宋体'

    __father_tk = None
    __tk_frame = None

    __select_file_bt = None  # 用于触发FileDialog.askopenfilename()的按钮
    __excel_path_entry = None  # 用于展示excel路径的Text控件
    # __excel_path_str = ''  # Excel文件路径

    __id_col_entry = None  # 身份证号码列
    __exe_col_entry = None  # 待执行操作的列
    __num_of_titles_entry = None  # 标题行数

    def __init__(self, father_tk, frame_type=FrameType.kUnknown, row=0, col=0):
        assert frame_type is not FrameType.kUnknown

        self.__father_tk = father_tk

        self.__tk_frame = self._create_tk_frame(frame_type, row, col)

        # 设置标题行
        # Label(self.__tk_frame, text='待拷贝文件' if frame_type is FrameType.kSrcFrame else '待复制文件',
        #       font=(self.__FONT, 16)).grid(row=0, column=0)
        Button(self.__tk_frame, text='点击选择待拷贝文件' if frame_type is FrameType.kSrcFrame else '点击选择待复制文件',
               command=self.__on_click_to_select_file).pack()

        # 设置excel文件相关属性
        # self.__select_file_bt = self._create_select_file_bt(self.__tk_frame)
        self.__excel_path_entry = self._create_lf_contains_entry(self.__tk_frame, 'Excel位置')
        # self.__excel_path_entry = self._create_lf_contains_text(self.__tk_frame, 'Excel位置', text_height=3)
        # self.__excel_path_str = ''

        # 设置excel中内容相关属性
        self.__id_col_entry = self._create_lf_contains_entry(self.__tk_frame, '身份证号码列')
        self.__exe_col_entry = self._create_lf_contains_entry(self.__tk_frame,
                                                              '待拷贝列' if frame_type is FrameType.kSrcFrame else '待复制列')
        self.__num_of_titles_entry = self._create_lf_contains_entry(self.__tk_frame, '标题行数')

    def _create_tk_frame(self, frame_type, row, col):
        color_str = kRED if frame_type is FrameType.kSrcFrame else kGREEN
        frame = Frame(self.__father_tk, padx=5, pady=5, bg=color_str)
        frame.grid(row=row, column=col, sticky='nsew', padx=5)
        return frame

    def __on_click_to_select_file(self):
        self.__excel_path_entry.delete(0, tkinter.END)
        path = askopenfilename()
        self.__excel_path_entry.insert(0, path)

    def _create_select_file_bt(self, father_tk):
        return Button(father_tk, text='点击选择文件', command=self.__on_click_to_select_file)

    def _create_lf_contains_entry(self, father_tk, title):
        lf = LabelFrame(father_tk, text=title, font=(self.__FONT, 12), padx=5, pady=5)  # LabelFrame与父窗口(frame)两端间隙各为5
        entry = Entry(lf)
        entry.pack()
        lf.pack()
        return entry

    def _create_lf_contains_text(self, father_tk, title, text_height=1):
        """创建指定LabelFrame对象

        :param father_tk: 父控件对象
        :type father_tk:

        :param title: LabelFrame标题
        :type title: string, e.g. 'Excel位置'

        :param text_height: LabelFrame中的Text高度
        :type text_height: int, e.g. 1

        :rtype: :class:`tkinter.Text`

        """
        lf = LabelFrame(father_tk, text=title, font=(self.__FONT, 12), padx=5, pady=5)  # LabelFrame与父窗口(frame)两端间隙各为5
        text = Text(lf, height=1)
        text.pack()
        lf.pack()
        return text

    def get_frame(self):
        return self.__tk_frame

    def get_excel_path(self):
        # return self.__excel_path_str
        return self.__excel_path_entry.get()

    def get_id_col(self):
        return self.__id_col_entry.get()

    def get_exe_col(self):
        return self.__exe_col_entry.get()

    def get_nums_of_title(self):
        # return self.__num_of_titles_text.get(1.0, tkinter.END + "-1c")
        return self.__num_of_titles_entry.get()


class BatchMatchAndCopyView:
    __father_tk = None
    __src_excel_wrapper = None
    __dst_excel_wrapper = None
    __start_bt = None

    def __init__(self, father_tk):
        self.__father_tk = father_tk
        self.__init_frames()
        self.__init_start_bt()

    def __init_frames(self):
        # 计算该frame的大小
        # win_width, win_height = get_tk_size(self.__father_tk)
        # width = win_width / 2 - 30  # 该view包含两个frame，两个frame之间、以及各自与父窗口之间，一共有3个间隙
        # height = 150
        self.__src_excel_wrapper = ExcelWrapper(self.__father_tk, FrameType.kSrcFrame, 0, 0)
        self.__dst_excel_wrapper = ExcelWrapper(self.__father_tk, FrameType.kDstFrame, 0, 1)

    def __init_start_bt(self):
        start_bt = Button(self.__father_tk, text='点击开始操作', command=self.__on_start_bt)
        start_bt.grid(row=1, column=0, sticky='nsew', pady=5, columnspan=2)  # columnspan表示该控件列数，最后的效果是合并单元格
        self.__start_bt = start_bt

    def __on_start_bt(self):
        # 执行过程中不允许再次点击
        self.__start_bt['state'] = tkinter.DISABLED

        src_wrapper = self.__src_excel_wrapper
        dst_wrapper = self.__dst_excel_wrapper

        try:
            # 加载excel文件
            src_excel = load_workbook(src_wrapper.get_excel_path())
            dst_excel = load_workbook(dst_wrapper.get_excel_path())

            # 获取activa的sheet
            src_active_sheet = src_excel.active
            dst_active_sheet = dst_excel.active

            # src_title_rows + 1 的原因是防止有多行标题，导致目标excel标题行被改动
            for src_row in range(int(src_wrapper.get_nums_of_title()) + 1, src_active_sheet.max_row + 1):
                src_id = (src_active_sheet[src_wrapper.get_id_col() + str(src_row)]).value
                for dst_row in range(int(dst_wrapper.get_nums_of_title()) + 1, dst_active_sheet.max_row + 1):
                    dst_id = (dst_active_sheet[dst_wrapper.get_id_col() + str(dst_row)]).value
                    # 对比身份证号码
                    if src_id == dst_id:
                        # 执行拷贝动作
                        dst_active_sheet[dst_wrapper.get_exe_col() + str(dst_row)].value = src_active_sheet[
                            src_wrapper.get_exe_col() + str(src_row)].value

            # 执行完成后，执行另存为操作，防止中途出错重写原文件
            dst_excel_path = dst_wrapper.get_excel_path()
            name_pos = dst_excel_path.rfind('/')
            path = dst_excel_path[0:name_pos]
            name = dst_excel_path[name_pos + 1:]
            dst_excel.save(path + '/new_' + name)
        except:
            tkinter.messagebox.showerror(title='错误', message='未知错误，请联系开发人员处理!')
        else:
            tkinter.messagebox.showinfo(title='信息', message='批量操作已完成!')

        # 当此任务完成后，enable button
        self.__start_bt['state'] = tkinter.NORMAL


def main():
    root_tk = Tk()
    root_tk.title('Excel批量操作工具')

    # WIDTH = 1000
    # HEIGHT = 500
    # root_tk.geometry(str(WIDTH) + 'x' + str(HEIGHT))

    view = BatchMatchAndCopyView(root_tk)

    root_tk.mainloop()


if __name__ == '__main__':
    main()
